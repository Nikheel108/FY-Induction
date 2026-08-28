"""
Attendance endpoints: public submission + admin management + CSV/Excel export.
"""

import csv
import io
import json
import logging
import uuid
from datetime import date, datetime, timedelta, timezone
from urllib.request import urlopen

from flask import Blueprint, jsonify, request, g, send_file
from openpyxl import Workbook
from sqlalchemy import or_, and_

from models import Student, Session, Attendance, EventSession, ValidPRN, Highlight
from services.database import db
from services.utils import admin_required

attendance_bp = Blueprint("attendance", __name__, url_prefix="/api")
logger = logging.getLogger(__name__)


# ----- Admin endpoints (protected) -----

@attendance_bp.route("/admin/attendance/session-stats", methods=["GET"])
@admin_required
def get_session_attendance_stats():
    """
    Get session-wise attendance tracking, total department student count,
    present student list, absent student list, and department summary breakdown.
    Query params: event_session_id (optional), department (optional).
    """
    session_id_arg = request.args.get("event_session_id")
    dept_arg = request.args.get("department", "").strip()

    target_session = None
    if session_id_arg:
        try:
            target_session = db.session.get(EventSession, int(session_id_arg))
        except ValueError:
            pass

    if not target_session:
        target_session = get_current_active_session()

    if not target_session:
        target_session = EventSession.query.order_by(EventSession.start_time.desc()).first()

    if not target_session:
        return jsonify({
            "success": True,
            "session": None,
            "stats": {
                "total_students": 0,
                "present_count": 0,
                "absent_count": 0,
                "attendance_percentage": 0,
                "by_department": []
            },
            "present_students": [],
            "absent_students": [],
            "highlights": []
        })

    # Master list of all students (registered + valid PRNs)
    all_students_map = {}
    
    # First, valid PRNs
    valid_prns = ValidPRN.query.all()
    for vp in valid_prns:
        all_students_map[vp.prn] = {
            "prn": vp.prn,
            "full_name": vp.expected_name or vp.prn,
            "department": vp.expected_department or "Computer Science and Engineering (AI & ML)",
            "student_email": "-",
            "student_phone": "-",
            "is_registered": False
        }
        
    # Second, registered students (override/enrich)
    students = Student.query.all()
    for s in students:
        all_students_map[s.prn] = {
            "prn": s.prn,
            "full_name": s.full_name or (all_students_map.get(s.prn, {}).get("full_name") or s.prn),
            "department": s.department or (all_students_map.get(s.prn, {}).get("department") or "Computer Science and Engineering (AI & ML)"),
            "student_email": s.student_email or "-",
            "student_phone": s.student_phone or "-",
            "is_registered": True
        }

    # Filter by department if specified
    if dept_arg:
        all_students_map = {prn: data for prn, data in all_students_map.items() if data["department"].lower() == dept_arg.lower()}

    # Fetch attendance records for this session
    att_records = Attendance.query.filter_by(event_session_id=target_session.id).all()
    present_dict = {att.prn: att for att in att_records}

    present_students = []
    absent_students = []

    for prn, info in all_students_map.items():
        if prn in present_dict:
            att = present_dict[prn]
            present_students.append({
                "prn": info["prn"],
                "full_name": info["full_name"],
                "department": info["department"],
                "student_email": info["student_email"],
                "student_phone": info["student_phone"],
                "is_registered": info["is_registered"],
                "attendance_id": att.id,
                "status": att.status,
                "date": att.date.isoformat(),
                "created_at": att.created_at.isoformat() + "Z",
                "ip": att.browser_session.ip_address if att.browser_session else "-",
                "location": att.browser_session.location if att.browser_session else "-"
            })
        else:
            absent_students.append(info)

    # Department breakdown stats
    dept_group = {}
    for prn, info in all_students_map.items():
        d_name = info["department"] or "Unknown"
        if d_name not in dept_group:
            dept_group[d_name] = {"total": 0, "present": 0, "absent": 0}
        dept_group[d_name]["total"] += 1
        if prn in present_dict:
            dept_group[d_name]["present"] += 1
        else:
            dept_group[d_name]["absent"] += 1

    by_department = []
    for d_name, d_counts in dept_group.items():
        pct = round((d_counts["present"] / d_counts["total"]) * 100, 1) if d_counts["total"] > 0 else 0
        by_department.append({
            "department": d_name,
            "total_students": d_counts["total"],
            "present_count": d_counts["present"],
            "absent_count": d_counts["absent"],
            "attendance_percentage": pct
        })

    total_studs = len(all_students_map)
    total_pres = len(present_students)
    total_abs = len(absent_students)
    overall_pct = round((total_pres / total_studs) * 100, 1) if total_studs > 0 else 0

    # Fetch highlights for this session
    session_highlights = Highlight.query.filter_by(event_session_id=target_session.id).order_by(Highlight.created_at.desc()).all()

    return jsonify({
        "success": True,
        "session": target_session.to_dict(),
        "stats": {
            "total_students": total_studs,
            "present_count": total_pres,
            "absent_count": total_abs,
            "attendance_percentage": overall_pct,
            "by_department": by_department
        },
        "present_students": present_students,
        "absent_students": absent_students,
        "highlights": [h.to_dict() for h in session_highlights]
    })


def get_location_from_ip(ip):
    """Fetch city and country from ip-api.com (free, no API key)."""
    try:
        url = f"http://ip-api.com/json/{ip}?fields=city,country"
        with urlopen(url, timeout=3) as resp:
            data = json.loads(resp.read().decode())
            if data.get("city") and data.get("country"):
                return f"{data['city']}, {data['country']}"
    except Exception as e:
        logger.warning("Geolocation failed for %s: %s", ip, e)
    return None


def get_or_create_session():
    """Retrieve session from cookie 'session_id' or create a new one."""
    session_id = request.cookies.get("session_id")
    session_obj = None

    if session_id:
        session_obj = Session.query.filter_by(session_id=session_id).first()

    if not session_obj:
        # Create new session
        new_id = str(uuid.uuid4())
        ip = request.headers.get("X-Forwarded-For", request.remote_addr)
        if ip and "," in ip:
            ip = ip.split(",")[0].strip()
        user_agent = request.headers.get("User-Agent", "")
        location = get_location_from_ip(ip) if ip else None

        session_obj = Session(
            session_id=new_id,
            ip_address=ip,
            location=location,
            user_agent=user_agent[:255]
        )
        db.session.add(session_obj)
        db.session.commit()
        g.new_session = True
    else:
        g.new_session = False

    g.session_obj = session_obj
    return session_obj


@attendance_bp.before_request
def attach_session():
    """Make session available in request context."""
    get_or_create_session()


@attendance_bp.after_request
def add_session_cookie(response):
    """If a new session was created, set the cookie."""
    if hasattr(g, 'new_session') and g.new_session:
        session_id = g.session_obj.session_id
        response.set_cookie('session_id', session_id, max_age=31536000,
                            httponly=True, samesite='Lax')
    return response


# ----- Event Sessions (Admin) -----

# ----- Event Sessions (Admin) -----

def parse_date_cell(val):
    import re
    if isinstance(val, (datetime, date)):
        return val.year, val.month, val.day
    if not val:
        return None
    val_str = str(val).strip()
    formats = [
        "%d-%b-%Y",  # 21-Sep-2026
        "%d-%b-%y",  # 21-Sep-26
        "%d/%m/%Y",  # 21/09/2026
        "%d/%m/%y",  # 21/09/26
        "%Y-%m-%d",  # 2026-09-21
        "%d %B %Y",  # 21 September 2026
        "%d %b %Y",  # 21 Sep 2026
    ]
    for fmt in formats:
        try:
            dt = datetime.strptime(val_str, fmt)
            return dt.year, dt.month, dt.day
        except ValueError:
            continue
    months_map = {
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
        'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
    }
    m = re.match(r"(\d+)[-\s/]+([a-zA-Z]+)[-\s/]+(\d+)", val_str)
    if m:
        try:
            d = int(m.group(1))
            m_name = m.group(2).lower()[:3]
            y = int(m.group(3))
            if y < 100:
                y += 2000
            if m_name in months_map:
                return y, months_map[m_name], d
        except Exception:
            pass
    return None

def parse_time_cell(time_val):
    import re
    if not time_val:
        return None, None
    time_str = str(time_val).strip()
    parts = re.split(r"[-–to]", time_str)
    if len(parts) < 2:
        return None, None
    
    def parse_part(p):
        p = p.strip().lower()
        is_pm = "pm" in p
        is_am = "am" in p
        p = p.replace("am", "").replace("pm", "").strip()
        p = p.replace(".", ":")
        subparts = p.split(":")
        h = int(subparts[0])
        m = int(subparts[1]) if len(subparts) > 1 else 0
        
        if is_pm and h < 12:
            h += 12
        elif is_am and h == 12:
            h = 0
        return h, m, is_pm, is_am
    
    try:
        h1, m1, pm1, am1 = parse_part(parts[0])
        h2, m2, pm2, am2 = parse_part(parts[1])
        
        if not pm1 and not am1:
            if h1 <= 7:
                h1 += 12
        if not pm2 and not am2:
            if h2 <= 7:
                h2 += 12
                
        if h2 < h1 and h1 < 12:
            if h2 <= 7:
                h2 += 12
        
        return (h1, m1), (h2, m2)
    except Exception:
        return None, None


@attendance_bp.route("/admin/event-sessions/upload-excel", methods=["POST"])
@admin_required
def upload_event_sessions_excel():
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "No selected file"}), 400

    if not file.filename.endswith(('.xls', '.xlsx')):
        return jsonify({"success": False, "message": "Unsupported file format. Please upload an Excel sheet."}), 400

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file)
        sheet = wb.active

        header_row_idx = -1
        col_mapping = {}
        
        # Scan first 20 rows to find headers
        for r_idx in range(1, 21):
            row_vals = [cell.value for cell in sheet[r_idx]]
            row_str_vals = [str(v).lower().strip() if v is not None else "" for v in row_vals]
            
            # Check if this row looks like a header (must contain at least 'time' and 'theme' or 'topic')
            if any("time" in v for v in row_str_vals) and (any("theme" in v for v in row_str_vals) or any("topic" in v for v in row_str_vals) or any("activity" in v for v in row_str_vals)):
                header_row_idx = r_idx
                for c_idx, val in enumerate(row_vals):
                    if not val:
                        continue
                    v_lower = str(val).lower().strip()
                    if "day" in v_lower:
                        col_mapping["day"] = c_idx
                    elif "date" in v_lower:
                        col_mapping["date"] = c_idx
                    elif "time" in v_lower:
                        col_mapping["time"] = c_idx
                    elif "theme" in v_lower or "ai & professional" in v_lower or "title" in v_lower:
                        col_mapping["theme"] = c_idx
                    elif "key topic" in v_lower:
                        col_mapping["key_topics"] = c_idx
                    elif "content" in v_lower:
                        col_mapping["content"] = c_idx
                    elif "student activity" in v_lower or ("activity" in v_lower and "student" in v_lower):
                        col_mapping["student_activity"] = c_idx
                    elif "ai tool" in v_lower:
                        col_mapping["ai_tools"] = c_idx
                    elif "interaction tool" in v_lower or "interaction" in v_lower:
                        col_mapping["interaction_tools"] = c_idx
                    elif "speaker" in v_lower or "faculty" in v_lower:
                        col_mapping["resource_speaker"] = c_idx
                break

        if header_row_idx == -1:
            col_mapping = {
                "day": 0,
                "date": 1,
                "time": 2,
                "theme": 3,
                "key_topics": 4,
                "content": 5,
                "student_activity": 6,
                "ai_tools": 7,
                "interaction_tools": 8,
                "resource_speaker": 9
            }
            header_row_idx = 1
            logger.warning("Header row not detected. Using fallback column mapping.")

        # Clear existing sessions if they don't have attendance
        try:
            EventSession.query.delete()
            db.session.commit()
        except Exception:
            db.session.rollback()
            return jsonify({"success": False, "message": "Cannot upload new schedule. Some existing sessions already have student attendance recorded."}), 400

        added_count = 0
        current_day = ""
        current_date_str = ""
        current_date_parsed = None

        for r_idx in range(header_row_idx + 1, sheet.max_row + 1):
            row = [cell.value for cell in sheet[r_idx]]
            if not row or all(v is None for v in row):
                continue

            raw_day = row[col_mapping["day"]] if "day" in col_mapping and col_mapping["day"] < len(row) else None
            raw_date = row[col_mapping["date"]] if "date" in col_mapping and col_mapping["date"] < len(row) else None
            raw_time = row[col_mapping["time"]] if "time" in col_mapping and col_mapping["time"] < len(row) else None
            raw_theme = row[col_mapping["theme"]] if "theme" in col_mapping and col_mapping["theme"] < len(row) else None
            raw_key_topics = row[col_mapping["key_topics"]] if "key_topics" in col_mapping and col_mapping["key_topics"] < len(row) else None
            raw_content = row[col_mapping["content"]] if "content" in col_mapping and col_mapping["content"] < len(row) else None
            raw_activity = row[col_mapping["student_activity"]] if "student_activity" in col_mapping and col_mapping["student_activity"] < len(row) else None
            raw_ai = row[col_mapping["ai_tools"]] if "ai_tools" in col_mapping and col_mapping["ai_tools"] < len(row) else None
            raw_interaction = row[col_mapping["interaction_tools"]] if "interaction_tools" in col_mapping and col_mapping["interaction_tools"] < len(row) else None
            raw_speaker = row[col_mapping["resource_speaker"]] if "resource_speaker" in col_mapping and col_mapping["resource_speaker"] < len(row) else None

            if not raw_theme and not raw_time:
                continue

            if raw_day is not None and str(raw_day).strip():
                current_day = str(raw_day).strip()
            if raw_date is not None:
                parsed_dt = parse_date_cell(raw_date)
                if parsed_dt:
                    current_date_parsed = parsed_dt
                    if isinstance(raw_date, (datetime, date)):
                        current_date_str = raw_date.strftime("%d-%b-%Y")
                    else:
                        current_date_str = str(raw_date).strip()
                elif str(raw_date).strip():
                    current_date_str = str(raw_date).strip()

            start_time_utc = None
            duration_minutes = 60

            if raw_time and current_date_parsed:
                start_hm, end_hm = parse_time_cell(raw_time)
                if start_hm and end_hm:
                    y, m, d = current_date_parsed
                    sh, sm = start_hm
                    eh, em = end_hm
                    
                    local_start = datetime(y, m, d, sh, sm)
                    start_time_utc = local_start - timedelta(hours=5, minutes=30)
                    
                    start_total_mins = sh * 60 + sm
                    end_total_mins = eh * 60 + em
                    if end_total_mins > start_total_mins:
                        duration_minutes = end_total_mins - start_total_mins
                    else:
                        duration_minutes = 60
            
            if not start_time_utc:
                if current_date_parsed:
                    y, m, d = current_date_parsed
                    start_time_utc = datetime(y, m, d, 9, 0) - timedelta(hours=5, minutes=30)
                else:
                    start_time_utc = datetime.utcnow()

            es = EventSession(
                title=str(raw_theme).strip() if raw_theme else "Induction Session",
                day=current_day,
                date_str=current_date_str,
                time_str=str(raw_time).strip() if raw_time else "-",
                theme=str(raw_theme).strip() if raw_theme else "-",
                key_topics=str(raw_key_topics).strip() if raw_key_topics else "-",
                content_to_be_covered=str(raw_content).strip() if raw_content else "-",
                student_activity=str(raw_activity).strip() if raw_activity else "-",
                ai_tools=str(raw_ai).strip() if raw_ai else "-",
                interaction_tools=str(raw_interaction).strip() if raw_interaction else "-",
                resource_speaker=str(raw_speaker).strip() if raw_speaker else "-",
                location="-",
                start_time=start_time_utc,
                duration_minutes=duration_minutes,
                attendance_limit_minutes=10
            )
            db.session.add(es)
            added_count += 1

        db.session.commit()
        return jsonify({"success": True, "message": f"Successfully imported {added_count} sessions.", "added": added_count})
    except Exception as e:
        logger.exception("Excel schedule import failed")
        return jsonify({"success": False, "message": f"Excel schedule import failed: {str(e)}"}), 500


@attendance_bp.route("/admin/event-sessions", methods=["POST"])
@admin_required
def create_event_session():
    """Create a new time-bound attendance session."""
    data = request.get_json() or {}
    title = data.get("title", "").strip()
    duration = data.get("duration_minutes") or 60
    attendance_limit = data.get("attendance_limit_minutes") or 10
    start_time_str = data.get("start_time")
    resource_speaker = data.get("resource_speaker", "-")
    location = data.get("location", "-")
    
    if not title or not start_time_str:
        return jsonify({"success": False, "message": "Missing required fields."}), 400
        
    try:
        start_time = datetime.fromisoformat(start_time_str.replace("Z", "+00:00")).astimezone(timezone.utc).replace(tzinfo=None)
    except ValueError:
        return jsonify({"success": False, "message": "Invalid duration or start time format."}), 400

    es = EventSession(
        title=title, 
        start_time=start_time, 
        duration_minutes=int(duration),
        attendance_limit_minutes=int(attendance_limit),
        resource_speaker=resource_speaker,
        location=location,
        day=data.get("day", "-"),
        date_str=data.get("date_str", "-"),
        time_str=data.get("time_str", "-"),
        theme=data.get("theme", "-"),
        key_topics=data.get("key_topics", "-"),
        content_to_be_covered=data.get("content_to_be_covered", "-"),
        student_activity=data.get("student_activity", "-"),
        ai_tools=data.get("ai_tools", "-"),
        interaction_tools=data.get("interaction_tools", "-"),
    )
    db.session.add(es)
    db.session.commit()
    return jsonify({"success": True, "message": "Session created.", "session": es.to_dict()}), 201


@attendance_bp.route("/admin/event-sessions", methods=["GET"])
@admin_required
def list_event_sessions():
    """List all event sessions."""
    sessions = EventSession.query.order_by(EventSession.start_time.desc()).all()
    return jsonify({"success": True, "sessions": [s.to_dict() for s in sessions]})


@attendance_bp.route("/admin/event-sessions/<int:session_id>", methods=["PUT"])
@admin_required
def edit_event_session(session_id):
    """Edit an existing event session."""
    es = db.session.get(EventSession, session_id)
    if not es:
        return jsonify({"success": False, "message": "Event session not found."}), 404

    data = request.get_json() or {}
    
    if "title" in data and data["title"].strip():
        es.title = data["title"].strip()
        
    if "duration_minutes" in data:
        try:
            es.duration_minutes = int(data["duration_minutes"])
        except ValueError:
            pass

    if "attendance_limit_minutes" in data:
        try:
            es.attendance_limit_minutes = int(data["attendance_limit_minutes"])
        except ValueError:
            pass

    if "start_time" in data and data["start_time"].strip():
        try:
            dt = datetime.fromisoformat(data["start_time"].replace("Z", "+00:00")).astimezone(timezone.utc).replace(tzinfo=None)
            es.start_time = dt
        except ValueError:
            pass
            
    if "resource_speaker" in data:
        es.resource_speaker = data["resource_speaker"].strip() or "-"
        
    if "location" in data:
        es.location = data["location"].strip() or "-"

    for field in ["day", "date_str", "time_str", "theme", "key_topics", "content_to_be_covered", "student_activity", "ai_tools", "interaction_tools"]:
        if field in data:
            es.__setattr__(field, data[field].strip() or "-")
        
    db.session.commit()
    return jsonify({"success": True, "message": "Session updated successfully.", "session": es.to_dict()})


@attendance_bp.route("/admin/event-sessions/<int:session_id>", methods=["DELETE"])
@admin_required
def delete_event_session(session_id):
    """Delete an event session."""
    es = db.session.get(EventSession, session_id)
    if not es:
        return jsonify({"success": False, "message": "Event session not found."}), 404
        
    try:
        # Check if attendance records exist for this session
        count = Attendance.query.filter_by(event_session_id=session_id).count()
        if count > 0:
            return jsonify({"success": False, "message": f"Cannot delete session. {count} attendance records are linked to it."}), 400
            
        db.session.delete(es)
        db.session.commit()
        return jsonify({"success": True, "message": "Event session deleted successfully."})
    except Exception as e:
        db.session.rollback()
        logger.exception("Failed to delete event session")
        return jsonify({"success": False, "message": "Database error while deleting."}), 500


@attendance_bp.route("/admin/event-sessions/clear", methods=["DELETE"])
@admin_required
def clear_all_event_sessions():
    """Delete all event sessions, highlights, and attendance records."""
    try:
        # Clear child records to prevent FK constraints
        Attendance.query.delete()
        Highlight.query.delete()
        EventSession.query.delete()
        db.session.commit()
        return jsonify({"success": True, "message": "All event sessions and linked records cleared successfully."})
    except Exception as e:
        db.session.rollback()
        logger.exception("Failed to clear event sessions")
        return jsonify({"success": False, "message": f"Database error: {str(e)}"}), 500





# ----- Public endpoint for active session -----

@attendance_bp.route("/attendance/active-session", methods=["GET"])
def get_active_session():
    """
    Returns the currently active event session.
    Active means current UTC time is between start_time and (start_time + attendance_limit + 5 mins).
    """
    now = datetime.utcnow()
    # Find the most recently started session that is still active
    # (In case of overlap, take the most recent one)
    sessions = EventSession.query.filter(EventSession.start_time <= now).order_by(EventSession.start_time.desc()).limit(5).all()
    
    for s in sessions:
        att_limit = s.attendance_limit_minutes if s.attendance_limit_minutes is not None else 15
        end_time = s.start_time + timedelta(minutes=att_limit + 5)
        if now <= end_time:
            already_marked = False
            
            # Check by browser session
            if hasattr(g, 'session_obj') and g.session_obj:
                existing = Attendance.query.filter_by(
                    session_id=g.session_obj.id,
                    event_session_id=s.id
                ).first()
                if existing:
                    already_marked = True
                    
            # Check by PRN if provided
            prn = request.args.get('prn', '').strip()
            if prn:
                existing_prn = Attendance.query.filter_by(
                    prn=prn,
                    event_session_id=s.id
                ).first()
                if existing_prn:
                    already_marked = True
                    
            return jsonify({
                "success": True, 
                "active_session": s.to_dict(),
                "already_marked": already_marked
            })
            
    return jsonify({"success": True, "active_session": None, "already_marked": False})


def get_current_active_session():
    """Helper to get active session on server-side"""
    now = datetime.utcnow()
    sessions = EventSession.query.filter(EventSession.start_time <= now).order_by(EventSession.start_time.desc()).limit(5).all()
    for s in sessions:
        att_limit = s.attendance_limit_minutes if s.attendance_limit_minutes is not None else 15
        end_time = s.start_time + timedelta(minutes=att_limit + 5)
        if now <= end_time:
            return s
    return None


def apply_filters(query):
    """Apply query filters from request args (reused by list and export)."""
    date_str = request.args.get("date")
    if date_str:
        try:
            filter_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            query = query.filter(Attendance.date == filter_date)
        except ValueError:
            pass

    prn = request.args.get("prn", "").strip()
    if prn:
        query = query.filter(Attendance.prn == prn)

    name = request.args.get("student_name", "").strip()
    if name:
        query = query.filter(Student.full_name.ilike(f"%{name}%"))
        
    event_session_id = request.args.get("event_session_id")
    if event_session_id:
        query = query.filter(Attendance.event_session_id == event_session_id)
        
    return query


# ----- Public endpoint for attendance submission -----
@attendance_bp.route("/attendance", methods=["POST"])
def submit_attendance():
    """
    Submit attendance for the currently active session.
    Requires PRN.
    """
    data = request.get_json() or {}
    prn = data.get("prn", "").strip()

    if not prn:
        return jsonify({"success": False, "message": "PRN is required."}), 400

    # Validate PRN exists
    student = Student.query.filter_by(prn=prn).first()
    if not student:
        return jsonify({"success": False, "message": "Invalid PRN. Student not found."}), 404

    # Ensure there is an active session
    active_session = get_current_active_session()
    if not active_session:
        return jsonify({"success": False, "message": "No active attendance session at the moment."}), 400

    # Ensure this student hasn't already marked attendance for this event session
    existing = Attendance.query.filter_by(
        prn=prn,
        event_session_id=active_session.id
    ).first()
    if existing:
        return jsonify({
            "success": False,
            "message": "You have already marked attendance for this session."
        }), 400

    # Ensure this browser session hasn't marked attendance for this event session yet
    existing_browser = Attendance.query.filter_by(
        session_id=g.session_obj.id,
        event_session_id=active_session.id
    ).first()
    if existing_browser:
        return jsonify({
            "success": False,
            "message": "This device has already been used to mark attendance for this session."
        }), 400

    # Create attendance record
    att = Attendance(
        prn=prn,
        session_id=g.session_obj.id,
        event_session_id=active_session.id,
        date=date.today(),
        status="present"
    )
    db.session.add(att)
    db.session.commit()

    return jsonify({
        "success": True,
        "message": "Attendance recorded successfully.",
        "attendance": {
            "id": att.id,
            "prn": att.prn,
            "event_session_title": active_session.title,
            "date": att.date.isoformat(),
            "time": att.created_at.isoformat(),
        }
    }), 201


# ----- Admin endpoints (protected) -----
@attendance_bp.route("/admin/attendance", methods=["GET"])
@admin_required
def list_attendance():
    """
    List attendance records with filters.
    Query params: date (YYYY-MM-DD), prn, student_name (search), page, per_page.
    """
    query = Attendance.query.join(Session).join(Student, Attendance.prn == Student.prn)
    query = apply_filters(query)

    # Pagination
    try:
        page = max(int(request.args.get("page", 1)), 1)
        per_page = min(max(int(request.args.get("per_page", 20)), 1), 100)
    except ValueError:
        page, per_page = 1, 20

    total = query.count()
    pagination = query.order_by(Attendance.date.desc(), Attendance.created_at.desc()) \
                     .paginate(page=page, per_page=per_page, error_out=False)

    # Build response with session details
    items = []
    for att in pagination.items:
        student = Student.query.filter_by(prn=att.prn).first()
        items.append({
            "id": att.id,
            "prn": att.prn,
            "student_name": student.full_name if student else att.prn,
            "department": student.department if student else "-",
            "date": att.date.isoformat(),
            "status": att.status,
            "event_session_title": att.event_session.title if att.event_session else "-",
            "session": {
                "ip": att.browser_session.ip_address if att.browser_session else "Unknown",
                "location": att.browser_session.location if att.browser_session else "Unknown",
                "user_agent": att.browser_session.user_agent if att.browser_session else "Unknown"
            }
        })

    return jsonify({
        "success": True,
        "message": "Attendance fetched successfully.",
        "data": {
            "items": items,
            "total": total,
            "page": page,
            "per_page": per_page,
            "pages": pagination.pages
        }
    })


@attendance_bp.route("/admin/attendance/<int:attendance_id>", methods=["DELETE"])
@admin_required
def delete_attendance(attendance_id):
    """Admin endpoint to demark (delete) an attendance record."""
    att = db.session.get(Attendance, attendance_id)
    if not att:
        return jsonify({"success": False, "message": "Attendance record not found."}), 404
        
    db.session.delete(att)
    db.session.commit()
    return jsonify({"success": True, "message": "Attendance record deleted successfully."})


@attendance_bp.route("/admin/attendance/mark", methods=["POST"])
@admin_required
def admin_mark_attendance():
    """Admin endpoint to manually mark attendance for a student."""
    data = request.get_json() or {}
    prn = data.get("prn", "").strip()
    event_session_id = data.get("event_session_id")

    if not prn or not event_session_id:
        return jsonify({"success": False, "message": "PRN and Event Session are required."}), 400

    student = Student.query.filter_by(prn=prn).first()
    if not student:
        return jsonify({"success": False, "message": f"Student with PRN {prn} not found."}), 404

    # Ensure this student hasn't already marked attendance for this event session
    existing = Attendance.query.filter_by(prn=prn, event_session_id=event_session_id).first()
    if existing:
        return jsonify({"success": False, "message": f"Student {prn} is already marked present for this session."}), 400

    # Create attendance record (using admin's session for audit)
    att = Attendance(
        prn=prn,
        session_id=g.session_obj.id,
        event_session_id=event_session_id,
        date=date.today(),
        status="present (manual)"
    )
    db.session.add(att)
    db.session.commit()

    return jsonify({"success": True, "message": f"Successfully marked {prn} as present."}), 201


@attendance_bp.route("/admin/attendance/export/csv", methods=["GET"])
@admin_required
def export_attendance_csv():
    """
    Export attendance records as CSV (respects filters).
    """
    query = Attendance.query.join(Session).join(Student, Attendance.prn == Student.prn)
    query = apply_filters(query)

    records = query.order_by(Attendance.date.desc(), Attendance.created_at.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    # Header
    writer.writerow(['Student Name', 'PRN', 'Department', 'Event Session', 'Date', 'Status', 'IP', 'Location', 'User Agent'])
    for att in records:
        student = Student.query.filter_by(prn=att.prn).first()
        writer.writerow([
            student.full_name if student else att.prn,
            att.prn,
            student.department if student else '-',
            att.event_session.title if att.event_session else '-',
            att.date.isoformat(),
            att.status,
            att.browser_session.ip_address if att.browser_session else '',
            (att.browser_session.location if att.browser_session else '') or '',
            (att.browser_session.user_agent if att.browser_session else '') or ''
        ])

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name='attendance_export.csv'
    )


@attendance_bp.route("/admin/attendance/export/excel", methods=["GET"])
@admin_required
def export_attendance_excel():
    """
    Export attendance records as Excel (.xlsx) (respects filters).
    """
    query = Attendance.query.join(Session).join(Student, Attendance.prn == Student.prn)
    query = apply_filters(query)

    records = query.order_by(Attendance.date.desc(), Attendance.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    # Headers
    headers = ['Student Name', 'PRN', 'Department', 'Event Session', 'Date', 'Status', 'IP', 'Location', 'User Agent']
    ws.append(headers)

    for att in records:
        student = Student.query.filter_by(prn=att.prn).first()
        ws.append([
            student.full_name if student else att.prn,
            att.prn,
            student.department if student else '-',
            att.event_session.title if att.event_session else '-',
            att.date.isoformat(),
            att.status,
            att.browser_session.ip_address if att.browser_session else '',
            (att.browser_session.location if att.browser_session else '') or '',
            (att.browser_session.user_agent if att.browser_session else '') or ''
        ])

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='attendance_export.xlsx'
    )